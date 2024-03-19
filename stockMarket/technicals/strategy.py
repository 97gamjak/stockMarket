import datetime as dt
import numpy as np
import pandas as pd
import yfinance as yf
import re
import glob
import shutil
import os
import matplotlib.pyplot as plt
import inspect

from decorator import decorator
from pathlib import Path
from beartype.typing import List, Optional
from tqdm import tqdm
from openpyxl import load_workbook
from finance_calendars import finance_calendars as fc

from .strategyObjects import StrategyObject, RuleEnum
from .trade.trade import Trade, TradeSettings
from .trade.enums import TradeOutcome
from stockMarket.utils import Period
from stockMarket.yfinance._common import adjust_price_data_from_df


@decorator
def finalize(func, *args, **kwargs):
    self = args[0]
    func(*args, **kwargs)
    for command in self.finalize_commands:
        os.system(command)


class StrategyFileSettings:
    def __init__(self,
                 base_path: str = "strategy_testing",
                 template_xlsx_path: Optional[str] = None,
                 template_xlsx_file: str = "template.xlsx",
                 xlsx_file: str = "screening.xlsx",
                 json_file: str = "strategy.json",
                 ) -> None:
        self.base_path = Path(base_path)
        self.template_xlsx_file = str(template_xlsx_path / template_xlsx_file)
        self.xlsx_filename = str(self.dir_path / xlsx_file)
        self.json_file = str(self.dir_path / json_file)

        if template_xlsx_path is not None:
            template_xlsx_path = Path(template_xlsx_path)
        else:
            template_xlsx_path = Path(__file__).parent / "templates"
        self.template_xlsx_file = str(template_xlsx_path / template_xlsx_file)
        self.xlsx_filename = str(self.dir_path / xlsx_file)

        self.json_file = str(self.dir_path / json_file)

    def build_dir_path(self, strategy_objects: List[StrategyObject]):

        strategy_names = [
            strategy_object.strategy_name for strategy_object in strategy_objects]
        return "_".join([name for name in sorted(strategy_names)])


class Strategy:
    storing_behavior_flags = (
        "full_overwrite", "matching_overwrite", "append", "abort", "numerical")

    default_file_options = {
        "base_path": "strategy_testing",
        "template_xlsx_path": None,
        "template_xlsx_file": "template.xlsx",
        "xlsx_file": "screening.xlsx",
        "json_file": "strategy.json"
    }

    def __init__(self,
                 strategy_objects: List[StrategyObject],
                 start_date: str,
                 end_date: str,
                 rule_enums: List[RuleEnum] = [RuleEnum.BULLISH],
                 trade_settings: Optional[TradeSettings] = None,
                 candle_period: str | Period | None = None,
                 storing_behavior: str | None = "numerical",
                 use_earnings_dates: bool = False,
                 finalize_commands=None,
                 **kwargs
                 ) -> None:

        self.setup_dates(start_date, end_date, candle_period)

        self.storing_behavior = storing_behavior.lower()

        self.strategy_objects = strategy_objects
        self.compile_strategy(strategy_objects, rule_enums)

        self.trade_settings = trade_settings if trade_settings is not None else TradeSettings()
        self.store_trade_settings()

        self.result_logger_filename = str(self.dir_path / "result_logger.txt")
        self.error_logger_filename = str(self.dir_path / "error_logger.txt")

        self.trades = {}
        self.meta_data_logger = {}
        self.error_logger = {}

        self.use_earnings_dates = use_earnings_dates

        if finalize_commands is None:
            finalize_commands = []

        self.finalize_commands = np.atleast_1d(finalize_commands)

    def setup_files(self, **kwargs):
        args_of_init = inspect.getfullargspec(
            StrategyFileSettings.__init__).args

        kwargs_for_file_settings = {}
        for key, value in kwargs.items():
            if key in args_of_init:
                kwargs_for_file_settings[key] = value

        file_settings = StrategyFileSettings(**kwargs_for_file_settings)
        self.base_path = file_settings.base_path
        self.template_xlsx_file = file_settings.template_xlsx_file
        self.xlsx_filename = file_settings.xlsx_filename
        self.json_file = file_settings.json_file
        self.dir_name = file_settings.build_dir_path(self.strategy_objects)

    def write_json(self, file_path: str) -> None:
        pass

    def compile_strategy(self,
                         strategy_objects: List[StrategyObject],
                         rule_enums: List[RuleEnum]
                         ) -> None:
        self.rule_enums = rule_enums

        strategy_names = [
            strategy_object.strategy_name for strategy_object in self.strategy_objects]
        self.dir_name = "_".join([name for name in sorted(strategy_names)])

        if self.storing_behavior not in self.storing_behavior_flags:
            raise NotImplementedError(
                f"Storing behavior not implemented, possible values are: {self.storing_behavior_flags}")

        if not self.base_path.exists():
            self.base_path.mkdir()

        self.dir_path = self.base_path / self.dir_name

        if self.storing_behavior == "numerical":

            if self.dir_path.exists():
                reserved_dir_names = glob.glob(
                    str(self.dir_path) + "_*")
                reserved_dir_names = [
                    name for name in reserved_dir_names if name.split("_")[-1].isdigit()]
                dir_numbers = sorted([int(name.split("_")[-1])
                                      for name in reserved_dir_names])
                if dir_numbers == []:
                    self.dir_path = Path(str(self.dir_path) + "_1")
                else:
                    self.dir_path = Path(
                        str(self.dir_path) + f"_{dir_numbers[-1] + 1}")

        elif self.storing_behavior == "abort":

            if self.dir_path.exists():
                raise FileExistsError(
                    f"Directory {self.dir_path} already exists, aborting")

        elif self.storing_behavior == "full_overwrite":
            if self.dir_path.exists():
                shutil.rmtree(self.dir_path)

        self.dir_path.mkdir()

    def store_trade_settings(self):
        self.trade_settings.write_to_json_file(
            str(self.dir_path / "trade_settings.json"))

    def setup_dates(self, start_date, end_date, candle_period):
        if candle_period is None:
            self.candle_period = Period('daily')
        else:
            self.candle_period = Period(candle_period)

        today = dt.datetime.now().date()
        self.start_date, self.end_date = _check_dates(start_date, end_date)

        period_time = (today - self.start_date).days + 1
        self.n_bars = np.ceil(
            period_time / self.candle_period.period_time.days)
        self.n_bars = int(self.n_bars) + 100

    def get_earnings_dates(self):
        self.earnings_calendar = {ticker: [] for ticker in self.tickers}
        date = pd.Timestamp(self.start_date).date()
        end_date = pd.Timestamp(self.end_date).date()
        while date < end_date:
            earnings = fc.get_earnings_by_date(date)
            for ticker in self.tickers:
                if ticker in earnings.index:
                    self.earnings_calendar[ticker].append(date)

            date += pd.Timedelta(days=1)

    def find_next_earnings_date(self, ticker, date):
        date = pd.Timestamp(date).date()
        earnings_dates = self.earnings_calendar[ticker]
        for earnings_date in earnings_dates:
            if earnings_date > date:
                return earnings_date, (earnings_date - date).days
        return "No earnings date found", None

    def write_xlsx_file(self):
        self.xlsx_file = load_workbook(filename=self.template_xlsx_file)
        xlsx_sheet = self.xlsx_file["stocks"]
        row = 3

        for trades in self.trades.values():
            for trade in trades:
                xlsx_sheet.cell(row=row, column=1).value = trade.TC_date
                xlsx_sheet.cell(row=row, column=2).value = trade.ENTRY_date
                xlsx_sheet.cell(row=row, column=3).value = trade.ticker
                xlsx_sheet.cell(row=row, column=4).value = trade.ENTRY
                xlsx_sheet.cell(
                    row=row, column=5).value = trade.R_ENTRY
                xlsx_sheet.cell(row=row, column=6).value = trade.SL
                xlsx_sheet.cell(
                    row=row, column=7).value = trade.TP_date
                xlsx_sheet.cell(row=row, column=8).value = trade.TP
                xlsx_sheet.cell(row=row, column=9).value = trade.EXIT_date
                xlsx_sheet.cell(row=row, column=10).value = trade.EXIT

                xlsx_sheet.cell(
                    row=row, column=11).value = trade.trade_status.value

                if self.use_earnings_dates:
                    xlsx_sheet.cell(row=row, column=12).value = self.find_next_earnings_date(
                        trade.ticker, trade.candle_date)[1]

                row += 1

        self.write_trade_settings_to_xlsx()

        self.xlsx_file.save(self.xlsx_filename)

    def write_trade_settings_to_xlsx(self):
        xlsx_sheet = self.xlsx_file["overview"]
        row = 1

        for key, value in self.trade_settings.to_json().items():
            xlsx_sheet.cell(row=row, column=1).value = key
            xlsx_sheet.cell(row=row, column=2).value = value

            row += 1

    @finalize
    def screen(self, tickers: List[str] | str) -> None:

        self.tickers = sorted(np.atleast_1d(tickers))

        if self.use_earnings_dates:
            self.get_earnings_dates()

        logger_file = open(self.result_logger_filename, "w")
        meta_file = open(str(self.dir_path / "meta_data_logger.txt"), "w")

        #fmt: off
        logger_file.write(f"Screening for {self.candle_period.period_string} candles from {self.start_date} to {self.end_date}\n\n\n")
        #fmt: on

        for ticker in tqdm(self.tickers):
            pricing_data = self.populate_pricing_data(ticker)
            self._screen_single_ticker(ticker, pricing_data)

        self.write_xlsx_file()

        logger_file.close()
        meta_file.close()

    def populate_pricing_data(self, ticker: str) -> None:
        error_file = open(self.error_logger_filename, "w")

        ticker = yf.Ticker(ticker)
        start_date = pd.Timestamp(self.start_date).date()
        start_date -= pd.Timedelta(days=100 *
                                   self.candle_period.period_time.days)

        pricing_data = ticker.history(
            auto_adjust=False,
            start=str(start_date),
            end=None,
            rounding=True,
            interval=self.candle_period.yf_interval
        )

        pricing_data = adjust_price_data_from_df(pricing_data)

        try:
            for strategy_object in self.strategy_objects:
                strategy_object.data = pricing_data
                strategy_object.calculate_indicators()
        except Exception as e:
            self.error_logger[ticker] = e
            error_file.write(f"Ticker: {ticker}\n")
            error_file.write(f"{self.error_logger[ticker]}\n")
            error_file.write("\n")
            error_file.flush()

            return None

        return pricing_data

    def _screen_single_ticker(self, ticker: str, pricing_data) -> None:

        if pricing_data is None:
            return

        self.trades[ticker] = []
        self.meta_data_logger[ticker] = {}

        end_index = _calculate_end_date_index(
            pricing_data, self.end_date)

        for index in -np.arange(end_index, len(pricing_data) + 1):
            date = get_date_from_pricing_data(pricing_data, index)

            if pd.Timestamp(self.start_date).date() > date:
                break

            date = pricing_data.iloc[index].name

            rule_outcome = [strategy_object.evaluate_rules(
                index) for strategy_object in self.strategy_objects]

            for rule_enum in self.rule_enums:

                rule_outcome = np.all([outcome[rule_enum.value]
                                      for outcome in rule_outcome])

                if rule_outcome:
                    trade = Trade(
                        ticker, pricing_data.iloc[index], self.trade_settings)

                    try:
                        trade.execute_trade(pricing_data)
                    except Exception as e:
                        print(f"Error executing trade for ticker {ticker}")
                        raise e

                    trade.condition = rule_enum.value
                    self.trades[ticker].append(trade)
                    for strategy_object in self.strategy_objects:
                        if hasattr(strategy_object, "meta_data"):
                            self.meta_data_logger[ticker][date] = strategy_object.meta_data[index]

    @finalize
    def plot_PL_histogramm(self):
        PL_win_values = []
        PL_loss_values = []
        for trades in self.trades.values():
            for trade in trades:
                if trade.outcome == TradeOutcome.WIN:
                    PL_win_values.append(trade.PL)
                elif trade.outcome == TradeOutcome.LOSS:
                    PL_loss_values.append(trade.PL)

        # Define the bins
        bins = np.arange(0, 6, 0.25)

        # Calculate the counts of 'win' and 'loss' trades in each bin
        win_counts, _ = np.histogram(PL_win_values, bins=bins)
        loss_counts, _ = np.histogram(PL_loss_values, bins=bins)

        # Calculate the PL ratio for each bin
        PL_ratio = win_counts / (win_counts + loss_counts)

        # Create a figure and a subplot
        fig, ax1 = plt.subplots()

        # Plot the histogram on ax1
        ax1.hist([PL_win_values, PL_loss_values], bins=bins, alpha=0.5,
                 label=['win', 'loss'], stacked=True, color=["g", "r"])
        ax1.legend(loc='upper left')

        # Create a second y-axis
        ax2 = ax1.twinx()

        # Plot the PL ratio on ax2
        ax2.plot(bins[:-1], PL_ratio, color='k', label='PL ratio')
        ax2.legend(loc='upper right')
        ax2.set_ylim(0, 1)

        plt.savefig(str(self.dir_path / "PL_histogramm.png"))


def _check_dates(start_date: str, end_date: str) -> tuple[dt.date, dt.date]:
    pattern = re.compile(r"\d{2}.\d{2}.\d{4}")
    if not pattern.fullmatch(start_date):
        raise ValueError("Start date is not in the correct format")
    if not pattern.fullmatch(end_date):
        raise ValueError("End date is not in the correct format")

    start_date = dt.datetime.strptime(start_date, "%d.%m.%Y").date()
    end_date = dt.datetime.strptime(end_date, "%d.%m.%Y").date()

    if end_date <= start_date:
        raise ValueError("End date is before or equal start date")

    return start_date, end_date


def _calculate_end_date_index(pricing_data, end_date):
    index = -1
    end_date = pd.Timestamp(end_date).date()
    while -index < len(pricing_data):
        date = get_date_from_pricing_data(pricing_data, index)

        if date < end_date:
            break

        index -= 1

    return abs(index)


def get_date_from_pricing_data(pricing_data, index):
    return pricing_data.index[index].date()
