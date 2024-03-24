import pandas as pd

from openpyxl import load_workbook
from beartype.typing import Dict, List, Optional

from .trade import Trade, TradeOutcome, TradeSettings
from stockMarket.utils.period import Period


class StrategyXLSXWriter:
    def __init__(self,
                 template_xlsx_filename: str,
                 xlsx_filename: str,
                 trade_settings: TradeSettings,
                 period: Period,
                 start_date: pd.Timestamp,
                 end_date: pd.Timestamp,
                 batch_size: pd.Timedelta,
                 ) -> None:

        self.template_xlsx_filename = template_xlsx_filename
        self.xlsx_filename = xlsx_filename
        self.trade_settings = trade_settings
        self.period = period
        self.start_date = start_date
        self.end_date = end_date
        self.batch_size = batch_size

    def write_xlsx_file(self,
                        trades: Dict[str, List[Trade]],
                        earnings_calendar: Dict[str, pd.Timestamp] = {}
                        ) -> None:

        start_date = self.start_date
        while True:
            end_date = start_date + self.batch_size

            # this way no only one fotal screening file is created
            if start_date == self.start_date and end_date > self.end_date:
                break

            self.write_single_xlsx_file(
                trades,
                earnings_calendar,
                start_date,
                end_date
            )

            start_date = end_date

            if start_date >= self.end_date:
                break

        self.write_single_xlsx_file(
            trades,
            earnings_calendar,
            filename=self.xlsx_filename.split(".")[0] + "_total.xlsx"
        )

    def write_single_xlsx_file(self,
                               trades: Dict[str, List[Trade]],
                               earnings_calendar: Dict[str, pd.Timestamp] = {},
                               start_date: pd.Timestamp = None,
                               end_date: pd.Timestamp = None,
                               filename: Optional[str] = None,
                               ) -> None:

        self.xlsx_file = load_workbook(self.template_xlsx_filename)

        self.write_all_possible_trades(
            trades,
            earnings_calendar,
            start_date,
            end_date
        )
        self.write_analytics_to_xlsx(
            trades,
            start_date,
            end_date
        )
        self.write_trade_settings_to_xlsx(
            start_date,
            end_date
        )

        if filename is None:
            xlsx_filename = self.xlsx_filename.split(".")[0] + \
                "_" + date_to_string(start_date) + "_"

            if end_date > self.end_date:
                xlsx_filename += date_to_string(self.end_date) + ".xlsx"
            else:
                xlsx_filename += date_to_string(end_date -
                                                pd.Timedelta(days=1)) + ".xlsx"
        else:
            xlsx_filename = filename

        self.xlsx_file.save(xlsx_filename)

    def write_all_possible_trades(self,
                                  trades: Dict[str, List[Trade]],
                                  earnings_calendar: Dict[str,
                                                          List[pd.Timestamp]] = {},
                                  start_date: pd.Timestamp = None,
                                  end_date: pd.Timestamp = None
                                  ) -> None:

        xlsx_sheet = self.xlsx_file["stocks"]
        headers = list(pd.read_excel(
            self.template_xlsx_filename, sheet_name="stocks").iloc[0])

        row = 3

        for trades in trades.values():
            for trade in trades:
                if not check_is_within_date_range(trade.TC_date, start_date, end_date):
                    continue

                TICKER_index = header_index(headers, "Ticker")

                TC_date_index = header_index(headers, "Entry Candle")
                ENTRY_date_index = header_index(headers, "Entry Date")

                ENTRY_index = header_index(headers, "Entry")
                R_ENTRY_index = header_index(headers, "Real Entry")

                SL_index = header_index(headers, "Stop Loss")
                TP_date_index = header_index(headers, "Target Date")
                TP_index = header_index(headers, "Target")

                EXIT_date_index = header_index(headers, "Exit Date")
                EXIT_index = header_index(headers, "Exit")

                STATUS = trade.trade_status.value
                STATUS_index = header_index(headers, "Status")

                EARNINGS_index = header_index(headers, "Days to Earnings")

                #fmt: off
                xlsx_sheet.cell(row=row, column=TICKER_index).value = trade.ticker
                xlsx_sheet.cell(row=row, column=TC_date_index).value = trade.TC_date
                xlsx_sheet.cell(row=row, column=ENTRY_date_index).value = trade.ENTRY_date
                xlsx_sheet.cell(row=row, column=ENTRY_index).value = trade.ENTRY
                xlsx_sheet.cell(row=row, column=R_ENTRY_index).value = trade.R_ENTRY

                xlsx_sheet.cell(row=row, column=SL_index).value = trade.SL
                xlsx_sheet.cell(row=row, column=TP_date_index).value = trade.TP_date
                xlsx_sheet.cell(row=row, column=TP_index).value = trade.TP
                xlsx_sheet.cell(row=row, column=EXIT_date_index).value = trade.EXIT_date
                xlsx_sheet.cell(row=row, column=EXIT_index).value = trade.EXIT

                xlsx_sheet.cell(row=row, column=STATUS_index).value = STATUS

                if earnings_calendar != {}:
                    _, days_until_earnings = self.find_next_earnings_date(
                        trade.candle_date,
                        earnings_calendar[trade.ticker]
                    )[1]

                    xlsx_sheet.cell(row=row, column=EARNINGS_index).value = days_until_earnings

                #fmt: on

                row += 1

    def find_next_earnings_date(self,
                                date: pd.Timestamp,
                                earnings_dates: List[pd.Timestamp],
                                ):
        date = pd.Timestamp(date).date()
        for earnings_date in earnings_dates:
            if earnings_date > date:
                return earnings_date, (earnings_date - date).days
        return "No earnings date found", None

    def write_analytics_to_xlsx(self,
                                trades: Dict[str, List[Trade]],
                                start_date: pd.Timestamp,
                                end_date: pd.Timestamp
                                ):

        xlsx_sheet = self.xlsx_file["analytics"]
        headers = list(pd.read_excel(
            self.template_xlsx_filename, sheet_name="analytics").iloc[0])

        row = 3

        for trades in trades.values():
            for trade in trades:
                if not check_is_within_date_range(trade.TC_date, start_date, end_date):
                    continue

                TICKER = trade.ticker
                TICKER_index = header_index(headers, "Ticker")

                xlsx_sheet.cell(row=row, column=TICKER_index).value = TICKER

                if trade.outcome_status != TradeOutcome.WIN and trade.outcome_status != TradeOutcome.LOSS:
                    row += 1
                    continue

                #fmt: off
                ENTRY_SL_index = header_index(headers, "Stop Loss")
                R_ENTRY_SL_index = header_index(headers, "Stop Loss Real")
                VOLATILITY_index = header_index(headers, "Volatility")

                TP_ENTRY_index = header_index(headers, "Target")
                TP_R_ENTRY_index = header_index(headers, "Target Real")

                PL_index = header_index(headers, "P/L")
                R_PL_index = header_index(headers, "P/L Real")

                WL = "W" if trade.outcome_status == TradeOutcome.WIN else "L"
                WL_index = header_index(headers, "W/L")

                SHARES_TO_BUY_index = header_index(headers, "Shares 2 Buy")
                PRED_INVEST_index = header_index(headers, "Predicted Investment")
                INVEST_index = header_index(headers, "Investment")
                DELTA_INVEST_index = header_index(headers, "Delta Investment")

                PRED_OUTCOME_index = header_index(headers, "Predicted Outcome")
                OUTCOME_index = header_index(headers, "Outcome")

                TOTAL_DAYS_index = header_index(headers, "Total Days")
                REQ_CAPITAL_index = header_index(headers, "Required Capital")


                xlsx_sheet.cell(row=row, column=ENTRY_SL_index).value = trade.ENTRY_SL
                xlsx_sheet.cell(row=row, column=R_ENTRY_SL_index).value = trade.R_ENTRY_SL
                xlsx_sheet.cell(row=row, column=TP_ENTRY_index).value = trade.TP_ENTRY
                xlsx_sheet.cell(row=row, column=TP_R_ENTRY_index).value = trade.TP_R_ENTRY
                xlsx_sheet.cell(row=row, column=PL_index).value = trade.PL
                xlsx_sheet.cell(row=row, column=R_PL_index).value = trade.R_PL

                xlsx_sheet.cell(row=row, column=WL_index).value = WL

                xlsx_sheet.cell(row=row, column=SHARES_TO_BUY_index).value = trade.SHARES_TO_BUY
                xlsx_sheet.cell(row=row, column=PRED_INVEST_index).value = trade.PRED_INVESTMENT
                xlsx_sheet.cell(row=row, column=INVEST_index).value = trade.INVESTMENT
                xlsx_sheet.cell(row=row, column=DELTA_INVEST_index).value = trade.DELTA_INVESTMENT

                xlsx_sheet.cell(row=row, column=PRED_OUTCOME_index).value = trade.PRED_OUTCOME
                xlsx_sheet.cell(row=row, column=OUTCOME_index).value = trade.OUTCOME
                xlsx_sheet.cell(row=row, column=TOTAL_DAYS_index).value = trade.TOTAL_DAYS
                xlsx_sheet.cell(row=row, column=REQ_CAPITAL_index).value = trade.REQ_CAPITAL

                xlsx_sheet.cell(row=row, column=VOLATILITY_index).value = trade.VOLATILITY

                #fmt: on
                row += 1

    def write_trade_settings_to_xlsx(self,
                                     start_date: pd.Timestamp = None,
                                     end_date: pd.Timestamp = None,
                                     ):
        xlsx_sheet = self.xlsx_file["overview"]

        if start_date is None:
            start_date = self.start_date
        if end_date is None:
            end_date = self.end_date

        xlsx_sheet.cell(row=1, column=1).value = "Start Date"
        xlsx_sheet.cell(row=1, column=2).value = start_date
        xlsx_sheet.cell(row=2, column=1).value = "End Date"
        xlsx_sheet.cell(row=2, column=2).value = end_date

        xlsx_sheet.cell(row=1, column=3).value = "candle_period"
        xlsx_sheet.cell(row=1, column=4).value = self.period.period_string

        row = 4

        for key, value in self.trade_settings.to_json().items():
            xlsx_sheet.cell(row=row, column=1).value = key
            xlsx_sheet.cell(row=row, column=2).value = value

            row += 1


def header_index(headers: List[str], header: str):
    headers = [header.lower().strip() for header in headers]
    return headers.index(header.lower()) + 1


def date_to_string(date: pd.Timestamp):
    return date.strftime("%d%m%Y")


def check_is_within_date_range(date: pd.Timestamp, start_date: pd.Timestamp, end_date: pd.Timestamp):
    if start_date is None:
        return True

    return start_date <= date < end_date
